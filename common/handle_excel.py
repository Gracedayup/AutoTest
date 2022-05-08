from common import project_path
import os
import openpyxl
class HandleExcel(object):
    def __init__(self, filename, sheetname=None):
        self.sheetname = sheetname
        self.excel_dir = project_path.testData_dir
        self.data_dir = os.path.join(self.excel_dir, filename)
        self.filename = self.data_dir
        print("文件地址：%s，sheet名：%s" % (self.excel_dir, self.data_dir))

    def read_excel(self):
        """
        从表格中读取内容
        :return:返回列表
        """
        # 加载工作簿
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        ws = wb[self.sheetname]
        # 获取表头信息
        data_list = []
        headers = tuple(ws.iter_rows(min_row=ws.min_row, max_row=ws.min_row, values_only=True))[0]
        print(headers)
        for content in tuple(ws.iter_rows(min_row=ws.min_row+1, values_only=True)):
            data_list.append(dict(zip(headers, content)))
            print(tuple(ws.columns))
        return data_list

    def write_excel(self, row, col, value):
        wb = openpyxl.load_workbook(self.filename)
        ws = wb.active
        ws = wb[self.sheetname]
        index = row+col
        ws[index] = value
        wb.save(self.filename)


if __name__ == '__main__':
    login_data = HandleExcel("login.xlsx", "login")
    print(login_data.read_excel())
    login_data.write_excel("G", "2", "pass2")


